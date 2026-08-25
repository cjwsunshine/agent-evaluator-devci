from app.models.models import db, TestCase, EvaluationSet, Agent, EvaluationTask, TaskTestCase, EvaluationResult
from app.utils.timeutil import iso_utc
import json
import os
import csv
from datetime import datetime
from typing import Dict, Any, List
from werkzeug.utils import secure_filename


class TestCaseService:
    UPLOAD_FOLDER = 'test_cases_uploads'
    ALLOWED_EXTENSIONS = {'json', 'csv', 'xlsx'}

    @staticmethod
    def _allowed_file(filename: str) -> bool:
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in TestCaseService.ALLOWED_EXTENSIONS

    @staticmethod
    def _resolve_agent_id(user_id, agent_id):
        if agent_id:
            return agent_id
        agent = Agent.query.filter_by(user_id=user_id, name='未绑定Agent', access_type='local').first()
        if not agent:
            agent = Agent(
                user_id=user_id,
                name='未绑定Agent',
                version='1.0.0',
                access_type='local',
                entry_function='run_agent',
                is_active=False
            )
            db.session.add(agent)
            db.session.flush()
        return agent.id

    @staticmethod
    def get_test_cases(user_id):
        try:
            test_cases = db.session.query(TestCase).filter_by(user_id=user_id).all()
            result = []
            for case in test_cases:
                result.append({
                    'id': case.id,
                    'name': case.name,
                    'agent_id': case.agent_id,
                    'agent_name': case.agent.name if case.agent else None,
                    'query': case.query,
                    'expected': case.expected,
                    'tags': case.tags,
                    'evaluation_tool': case.evaluation_tool,
                    'metric': case.metric,
                    'input_payload': case.input_payload,
                    'expected_payload': case.expected_payload,
                    'created_at': iso_utc(case.created_at)
                })
            return {'success': True, 'data': result}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    @staticmethod
    def create_test_case(user_id, data):
        try:
            test_case = TestCase(
                user_id=user_id,
                agent_id=data.get('agent_id'),
                set_id=data.get('set_id'),
                name=data.get('name'),
                query=data.get('query'),
                expected=data.get('expected'),
                tags=data.get('tags'),
                evaluation_tool=data.get('evaluation_tool', 'deepeval'),
                metric=data.get('metric'),
                input_payload=data.get('input_payload'),
                expected_payload=data.get('expected_payload')
            )
            db.session.add(test_case)
            db.session.commit()
            return {'success': True, 'message': '测试用例创建成功', 'data': {'id': test_case.id}}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def update_test_case(user_id, case_id, data):
        try:
            test_case = db.session.query(TestCase).filter_by(id=case_id, user_id=user_id).first()
            if not test_case:
                return {'success': False, 'message': '测试用例不存在'}

            test_case.agent_id = data.get('agent_id', test_case.agent_id)
            test_case.set_id = data.get('set_id', test_case.set_id)
            test_case.name = data.get('name', test_case.name)
            test_case.query = data.get('query', test_case.query)
            test_case.expected = data.get('expected', test_case.expected)
            test_case.tags = data.get('tags', test_case.tags)
            test_case.evaluation_tool = data.get('evaluation_tool', test_case.evaluation_tool)
            test_case.metric = data.get('metric', test_case.metric)
            if 'input_payload' in data:
                test_case.input_payload = data['input_payload']
            if 'expected_payload' in data:
                test_case.expected_payload = data['expected_payload']

            db.session.commit()
            return {'success': True, 'message': '测试用例更新成功'}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def delete_test_case(user_id, case_id):
        try:
            test_case = db.session.query(TestCase).filter_by(id=case_id, user_id=user_id).first()
            if not test_case:
                return {'success': False, 'message': '测试用例不存在'}

            task_cases = TaskTestCase.query.filter_by(test_case_id=case_id).all()
            for task_case in task_cases:
                EvaluationResult.query.filter_by(task_case_id=task_case.id).delete()
                db.session.delete(task_case)
            db.session.delete(test_case)
            db.session.commit()
            return {'success': True, 'message': '测试用例删除成功'}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def upload_test_cases_file(user_id, file) -> Dict[str, Any]:
        """上传测试用例文件"""
        try:
            if not file or not TestCaseService._allowed_file(file.filename):
                return {'success': False, 'message': '不支持的文件类型，仅支持.json、.csv和.xlsx文件'}

            # 确保上传目录存在
            upload_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                TestCaseService.UPLOAD_FOLDER
            )
            os.makedirs(upload_dir, exist_ok=True)

            filename = secure_filename(f"user_{user_id}_{file.filename}")
            filepath = os.path.join(upload_dir, filename)
            file.save(filepath)

            # 解析文件预览
            test_cases = TestCaseService._parse_file(filepath)

            return {
                'success': True,
                'message': '文件上传成功',
                'data': {
                    'filename': filename,
                    'test_cases': test_cases,
                    'total_count': len(test_cases)
                }
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    @staticmethod
    def _parse_file(filepath: str) -> List[Dict[str, Any]]:
        """解析测试用例文件"""
        if filepath.endswith('.json'):
            return TestCaseService._parse_json_file(filepath)
        elif filepath.endswith('.csv'):
            return TestCaseService._parse_csv_file(filepath)
        elif filepath.endswith('.xlsx'):
            return TestCaseService._parse_excel_file(filepath)
        return []

    @staticmethod
    def _parse_json_file(filepath: str) -> List[Dict[str, Any]]:
        """解析JSON格式的测试用例文件"""
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)

        test_cases = []

        # 支持多种JSON格式
        if isinstance(data, list):
            # 直接是测试用例数组
            for i, item in enumerate(data):
                test_case = TestCaseService._extract_test_case(item, i + 1)
                if test_case:
                    test_cases.append(test_case)
        elif isinstance(data, dict) and 'test_cases' in data:
            # 包含test_cases字段
            for i, item in enumerate(data['test_cases']):
                test_case = TestCaseService._extract_test_case(item, i + 1)
                if test_case:
                    test_case['evaluation_tool'] = item.get('evaluation_tool') or data.get('evaluation_tool') or 'deepeval'
                    test_case['metric'] = TestCaseService._normalize_metric(item.get('metric') or data.get('metric'))
                    test_cases.append(test_case)
        elif isinstance(data, dict):
            # 单个测试用例
            test_case = TestCaseService._extract_test_case(data, 1)
            if test_case:
                test_cases.append(test_case)

        return test_cases

    @staticmethod
    def _parse_csv_file(filepath: str) -> List[Dict[str, Any]]:
        """解析CSV格式的测试用例文件"""
        test_cases = []
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader):
                test_case = TestCaseService._row_to_test_case(row, i)
                if test_case:
                    test_cases.append(test_case)

        return test_cases

    @staticmethod
    def _parse_excel_file(filepath: str) -> List[Dict[str, Any]]:
        """解析Excel格式(.xlsx)的测试用例文件：首行为表头，每行一个用例，每列一个参数"""
        from openpyxl import load_workbook

        test_cases = []
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = None
            data_index = 0
            for raw in rows:
                if headers is None:
                    # 首个非空行作为表头
                    if raw is None or all(cell is None or str(cell).strip() == '' for cell in raw):
                        continue
                    headers = [str(cell).strip() if cell is not None else '' for cell in raw]
                    continue
                if raw is None or all(cell is None or str(cell).strip() == '' for cell in raw):
                    continue
                row = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = raw[idx] if idx < len(raw) else None
                    row[header] = '' if value is None else str(value).strip()
                test_case = TestCaseService._row_to_test_case(row, data_index)
                if test_case:
                    test_cases.append(test_case)
                data_index += 1
        finally:
            workbook.close()

        return test_cases

    @staticmethod
    def build_excel_sample(rows: List[Dict[str, Any]]) -> bytes:
        """根据样例行生成 Excel(.xlsx) 模板：表头为字段名，每行一个用例。返回文件字节。"""
        from openpyxl import Workbook
        import io

        rows = rows or []
        # 收集所有出现过的列，保持常用字段在前的顺序。
        # 注意：Agent / 评测工具 / 评测阶段 / 评测指标由前台选择统一注入，不写入模板。
        preferred = ['name', 'query', 'expected', 'tags', 'context', 'reference',
                     'ground_truth', 'input_payload', 'expected_payload']
        seen = []
        for row in rows:
            for key in row.keys():
                if key not in seen:
                    seen.append(key)
        headers = [key for key in preferred if key in seen] + [key for key in seen if key not in preferred]
        if not headers:
            headers = preferred

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'test_cases'
        sheet.append(headers)
        for row in rows:
            values = []
            for key in headers:
                value = row.get(key, '')
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                values.append('' if value is None else value)
            sheet.append(values)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _row_to_test_case(row: Dict[str, Any], index: int = 0) -> Dict[str, Any]:
        """将一行(CSV/Excel)按列名映射为测试用例，兼容多种表头写法"""
        def pick(*keys, default=''):
            for key in keys:
                if key in row and row[key] not in (None, ''):
                    return row[key]
            return default

        item = {
            'name': pick('name', 'Name', '测试用例名称', default=f'测试用例_{index+1}'),
            'query': pick('query', 'Query', '输入'),
            'expected': pick('expected', 'Expected', '预期输出'),
            'tags': pick('tags', 'Tags', '标签'),
            'agent_id': pick('agent_id', 'Agent ID', 'agent', 'Agent', default=None) or None,
            'evaluation_tool': pick('evaluation_tool', 'tool', '评测工具', default='deepeval'),
            'metric': pick('metric', '指标'),
            'context': pick('context', 'contexts', 'retrieved_contexts', '上下文', default=None) or None,
            'reference': pick('reference', 'ground_truth', '参考答案', default=None) or None,
            'ground_truth': pick('ground_truth', '标准答案', default=None) or None,
            'input_payload': pick('input_payload', default=None) or None,
            'expected_payload': pick('expected_payload', default=None) or None
        }
        input_payload, expected_payload = TestCaseService._build_payloads(item)
        test_case = {**item, 'input_payload': input_payload, 'expected_payload': expected_payload}
        return test_case if test_case['query'] else None

    @staticmethod
    def _normalize_metric(metric: str) -> str:
        metric_map = {
            'AnswerRelevancyMetric': 'answer_relevancy',
            'FaithfulnessMetric': 'factuality',
            'ToolCorrectnessMetric': 'tool_call_accuracy',
            'TaskCompletionMetric': 'task_completion',
            'GoalAccuracyMetric': 'goal_accuracy',
            'StepEfficiencyMetric': 'step_efficiency',
            '对抗鲁棒性': 'adversarial_robustness',
            '内容安全拦截率': 'content_safety_interception',
            '红队测试通过率': 'red_team_pass_rate',
            '内容安全拦截率（红队）': 'content_safety_interception_redteam'
        }
        return metric_map.get(metric, metric or '')

    @staticmethod
    def _parse_json_value(value):
        if not isinstance(value, str):
            return value
        text = value.strip()
        if not text:
            return None
        if not ((text.startswith('{') and text.endswith('}')) or (text.startswith('[') and text.endswith(']'))):
            return value
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return value

    @staticmethod
    def _compact_payload(payload):
        if not isinstance(payload, dict):
            return payload
        return {key: value for key, value in payload.items() if value not in (None, '', [])} or None

    @staticmethod
    def _build_payloads(item: Dict[str, Any]):
        input_payload = TestCaseService._parse_json_value(item.get('input_payload'))
        expected_payload = TestCaseService._parse_json_value(item.get('expected_payload'))
        input_payload = input_payload if isinstance(input_payload, dict) else {}
        expected_payload = expected_payload if isinstance(expected_payload, dict) else {}

        contexts = item.get('contexts') or item.get('context') or item.get('retrieved_contexts')
        contexts = TestCaseService._parse_json_value(contexts)
        if contexts and 'contexts' not in input_payload:
            input_payload['contexts'] = contexts if isinstance(contexts, list) else [contexts]

        reference = item.get('reference') or item.get('ground_truth')
        if reference and 'reference' not in expected_payload:
            expected_payload['reference'] = reference
        if item.get('ground_truth') and 'ground_truth' not in expected_payload:
            expected_payload['ground_truth'] = item.get('ground_truth')

        return TestCaseService._compact_payload(input_payload), TestCaseService._compact_payload(expected_payload)

    @staticmethod
    def _extract_test_case(item: Dict[str, Any], index: int) -> Dict[str, Any]:
        """从字典中提取测试用例字段"""
        if not isinstance(item, dict):
            return None

        input_payload, expected_payload = TestCaseService._build_payloads(item)
        expected = (
            item.get('expected') or item.get('expected_output') or item.get('answer')
            or item.get('reference') or item.get('ground_truth')
            or (expected_payload or {}).get('reference') or (expected_payload or {}).get('ground_truth')
            or ''
        )
        return {
            'name': item.get('name') or item.get('description', f'测试用例_{index}'),
            'query': item.get('query') or item.get('input') or item.get('question', ''),
            'expected': expected,
            'tags': item.get('tags') or '',
            'agent_id': item.get('agent_id'),
            'evaluation_tool': item.get('evaluation_tool') or item.get('tool') or 'deepeval',
            'metric': TestCaseService._normalize_metric(item.get('metric')),
            'input_payload': input_payload,
            'expected_payload': expected_payload
        }

    @staticmethod
    def confirm_import(user_id, filename: str, selected_indices: List[int] = None, agent_id: int = None, evaluation_tool: str = 'deepeval', metric: str = None) -> Dict[str, Any]:
        """确认并导入测试用例到数据库"""
        try:
            filepath = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                TestCaseService.UPLOAD_FOLDER,
                filename
            )

            if not os.path.exists(filepath):
                return {'success': False, 'message': '文件不存在'}

            test_cases = TestCaseService._parse_file(filepath)

            # 过滤选中的测试用例
            if selected_indices is not None:
                test_cases = [tc for i, tc in enumerate(test_cases) if i in selected_indices]

            # 验证测试用例
            validation_errors = TestCaseService._validate_test_cases(test_cases)
            if validation_errors:
                return {
                    'success': False,
                    'message': '测试用例验证失败',
                    'errors': validation_errors
                }

            # 批量导入
            imported_count = 0
            for tc in test_cases:
                test_case = TestCase(
                    user_id=user_id,
                    agent_id=agent_id or tc.get('agent_id'),
                    name=tc['name'],
                    query=tc['query'],
                    expected=tc['expected'],
                    tags=tc.get('tags', ''),
                    evaluation_tool=evaluation_tool or tc.get('evaluation_tool', 'deepeval'),
                    metric=metric or tc.get('metric'),
                    input_payload=tc.get('input_payload'),
                    expected_payload=tc.get('expected_payload')
                )
                db.session.add(test_case)
                imported_count += 1

            db.session.commit()

            return {
                'success': True,
                'message': f'成功导入{imported_count}个测试用例',
                'data': {'imported_count': imported_count}
            }
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def _validate_test_cases(test_cases: List[Dict[str, Any]]) -> List[str]:
        """验证测试用例格式正确性"""
        errors = []
        for i, tc in enumerate(test_cases):
            if not tc.get('query'):
                errors.append(f"第{i+1}个测试用例缺少输入(query)字段")
            if not tc.get('expected'):
                errors.append(f"第{i+1}个测试用例缺少预期输出(expected)字段")
        return errors

    @staticmethod
    def batch_import(user_id, test_cases: List[Dict[str, Any]]) -> Dict[str, Any]:
        """批量导入测试用例"""
        try:
            # 验证测试用例
            validation_errors = TestCaseService._validate_test_cases(test_cases)
            if validation_errors:
                return {
                    'success': False,
                    'message': '测试用例验证失败',
                    'errors': validation_errors
                }

            # 批量导入
            imported_ids = []
            for tc in test_cases:
                test_case = TestCase(
                    user_id=user_id,
                    agent_id=tc.get('agent_id'),
                    set_id=tc.get('set_id'),
                    name=tc.get('name', f'测试用例_{len(imported_ids)+1}'),
                    query=tc['query'],
                    expected=tc.get('expected', ''),
                    tags=tc.get('tags', ''),
                    evaluation_tool=tc.get('evaluation_tool', 'deepeval'),
                    metric=tc.get('metric'),
                    input_payload=tc.get('input_payload'),
                    expected_payload=tc.get('expected_payload')
                )
                db.session.add(test_case)
                db.session.flush()
                imported_ids.append(test_case.id)

            db.session.commit()

            return {
                'success': True,
                'message': f'成功导入{len(imported_ids)}个测试用例',
                'data': {'imported_ids': imported_ids, 'count': len(imported_ids)}
            }
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def _get_set_metric(cases):
        metrics = sorted({case.metric for case in cases if case.metric})
        return metrics[0] if len(metrics) == 1 else ','.join(metrics)

    @staticmethod
    def _get_cases_status(cases, fallback='pending'):
        case_ids = [case.id for case in cases]
        if not case_ids:
            return fallback

        latest_task = (
            db.session.query(EvaluationTask)
            .join(TaskTestCase, TaskTestCase.task_id == EvaluationTask.id)
            .filter(TaskTestCase.test_case_id.in_(case_ids))
            .order_by(EvaluationTask.created_at.desc(), EvaluationTask.id.desc())
            .first()
        )
        return latest_task.status if latest_task else fallback

    @staticmethod
    def _get_cases_last_run(cases):
        """返回这批用例最近一次执行的时间（end_time 优先，其次 start_time），无则 None。"""
        case_ids = [case.id for case in cases]
        if not case_ids:
            return None
        latest_task = (
            db.session.query(EvaluationTask)
            .join(TaskTestCase, TaskTestCase.task_id == EvaluationTask.id)
            .filter(TaskTestCase.test_case_id.in_(case_ids))
            .order_by(EvaluationTask.created_at.desc(), EvaluationTask.id.desc())
            .first()
        )
        if not latest_task:
            return None
        return latest_task.end_time or latest_task.start_time

    @staticmethod
    def _get_set_updated_at(cases):
        """评测集"更新时间"：内容最近修改时间与最近一次执行时间中较晚者。"""
        candidates = [case.updated_at for case in cases if case.updated_at]
        last_run = TestCaseService._get_cases_last_run(cases)
        if last_run:
            candidates.append(last_run)
        return max(candidates) if candidates else None

    @staticmethod
    def get_evaluation_sets(user_id):
        try:
            evaluation_sets = EvaluationSet.query.filter_by(user_id=user_id).order_by(EvaluationSet.created_at.desc()).all()
            result = []
            for evaluation_set in evaluation_sets:
                metric = TestCaseService._get_set_metric(evaluation_set.test_cases)
                result.append({
                    'id': evaluation_set.id,
                    'name': metric if '默认评测集' in evaluation_set.name and metric else evaluation_set.name,
                    'agent_id': evaluation_set.agent_id,
                    'agent_name': evaluation_set.agent.name if evaluation_set.agent else None,
                    'evaluation_tool': evaluation_set.evaluation_tool,
                    'metric': metric,
                    'status': TestCaseService._get_cases_status(evaluation_set.test_cases, evaluation_set.status),
                    'test_case_count': len(evaluation_set.test_cases),
                    'created_at': iso_utc(evaluation_set.created_at),
                    'updated_at': iso_utc(TestCaseService._get_set_updated_at(evaluation_set.test_cases) or evaluation_set.created_at),
                    'test_cases': [{
                        'id': case.id,
                        'name': case.name,
                        'query': case.query,
                        'expected': case.expected,
                        'tags': case.tags,
                        'metric': case.metric,
                        'input_payload': case.input_payload,
                        'expected_payload': case.expected_payload
                    } for case in evaluation_set.test_cases]
                })

            orphan_cases = db.session.query(TestCase).filter_by(user_id=user_id, set_id=None).all()
            grouped_cases = {}
            for case in orphan_cases:
                key = (case.agent_id, case.evaluation_tool or 'deepeval', case.metric or '')
                grouped_cases.setdefault(key, []).append(case)

            for (agent_id, evaluation_tool, metric), cases in grouped_cases.items():
                agent_name = cases[0].agent.name if cases[0].agent else '未绑定Agent'
                metric_name = metric or '未设置指标'
                result.append({
                    'id': f'orphan-{agent_id or "none"}-{evaluation_tool}::{metric or "none"}',
                    'name': metric_name,
                    'agent_id': agent_id,
                    'agent_name': agent_name,
                    'evaluation_tool': evaluation_tool,
                    'metric': metric,
                    'status': TestCaseService._get_cases_status(cases),
                    'test_case_count': len(cases),
                    'created_at': iso_utc(min(case.created_at for case in cases)),
                    'updated_at': iso_utc(TestCaseService._get_set_updated_at(cases) or min(case.created_at for case in cases)),
                    'test_cases': [{
                        'id': case.id,
                        'name': case.name,
                        'query': case.query,
                        'expected': case.expected,
                        'tags': case.tags,
                        'metric': case.metric,
                        'input_payload': case.input_payload,
                        'expected_payload': case.expected_payload
                    } for case in cases]
                })
            # 按更新时间由新到旧排序（updated_at 为带时区的 ISO 字符串，可直接字符串比较）
            result.sort(key=lambda item: item.get('updated_at') or item.get('created_at') or '', reverse=True)
            return {'success': True, 'data': result}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    # 每个指标所需的 payload 字段。缺失时保存评测集会被拒绝，并提示用户如何补齐。
    # 结构：{(evaluation_tool, metric): [(path, hint), ...]}
    # path 用 'input_payload.contexts' / 'expected_payload.reference' 表达。
    METRIC_REQUIRED_FIELDS = {
        # RAGAS
        ('ragas', 'answer_correctness'): [('expected_payload.reference', '参考答案')],
        ('ragas', 'faithfulness'): [('input_payload.contexts', '检索/工具返回的上下文')],
        ('ragas', 'context_precision'): [('input_payload.contexts', '检索/工具返回的上下文'), ('expected_payload.reference', '参考答案')],
        ('ragas', 'context_recall'): [('input_payload.contexts', '检索/工具返回的上下文'), ('expected_payload.reference', '参考答案')],
        ('ragas', 'context_entity_recall'): [('input_payload.contexts', '检索/工具返回的上下文'), ('expected_payload.reference', '参考答案')],
        ('ragas', 'noise_sensitivity'): [('input_payload.contexts', '检索/工具返回的上下文'), ('expected_payload.reference', '参考答案')],
        # TruLens
        ('trulens', 'groundedness'): [('input_payload.context', '检索/工具返回的原始上下文（groundedness 需要用来核对回答的每句话是否有依据）')],
        ('trulens', 'context_relevance'): [('input_payload.context', '检索/工具返回的原始上下文')],
        # DeepEval —— tool_correctness 强要求两侧工具调用列表都存在（可为空列表，但字段必须有）
        ('deepeval', 'tool_correctness'): [
            ('expected_payload.expected_tool_calls', '期望的工具调用列表（可为空数组 []，用于对比 agent 的实际调用）'),
        ],
    }

    @staticmethod
    def _get_by_path(item: Dict[str, Any], path: str):
        """从测试项对象取 'input_payload.contexts' 这类嵌套字段。用例既支持嵌套 payload，
        也支持扁平字段（如 top-level 的 'context' / 'reference' / 'ground_truth'），
        因此校验时按后端 _build_payloads 的合并规则一并识别。"""
        top, key = path.split('.', 1) if '.' in path else (path, None)
        value = item.get(top)
        if isinstance(value, str):
            try:
                value = json.loads(value)
            except Exception:
                pass
        if key:
            # 优先嵌套 payload 里的字段
            if isinstance(value, dict) and value.get(key) not in (None, '', []):
                return value[key]
            # 兼容扁平字段：contexts 可写成 context / retrieved_contexts；reference 可写成 ground_truth
            flat_aliases = {
                'contexts': ['contexts', 'context', 'retrieved_contexts'],
                'context': ['context', 'contexts', 'retrieved_contexts'],
                'reference': ['reference', 'ground_truth'],
                'ground_truth': ['ground_truth', 'reference'],
            }.get(key, [key])
            for alias in flat_aliases:
                v = item.get(alias)
                if v not in (None, '', []):
                    return v
            return None
        return value if value not in (None, '', []) else None

    @staticmethod
    def validate_test_cases_for_metric(test_cases: List[Dict[str, Any]], evaluation_tool: str, metric: str) -> List[str]:
        """按所选（工具, 指标）校验每条用例是否具备必需字段，返回可读的错误列表；空列表表示通过。"""
        errors = []
        rules = TestCaseService.METRIC_REQUIRED_FIELDS.get((evaluation_tool, metric), [])
        for idx, tc in enumerate(test_cases or []):
            if not isinstance(tc, dict):
                continue
            name = tc.get('name') or f'第 {idx + 1} 条'
            # 基础字段
            query_value = tc.get('query')
            if not query_value or (isinstance(query_value, str) and not query_value.strip()):
                errors.append(f'用例「{name}」缺少 query（查询内容）')
            # 指标专属字段
            for path, hint in rules:
                if TestCaseService._get_by_path(tc, path) is None:
                    errors.append(f'用例「{name}」缺少 {path}（{hint}）——{metric} 指标要求此字段')
        return errors

    @staticmethod
    def create_evaluation_set(user_id, data):
        try:
            evaluation_tool = data.get('evaluation_tool', 'deepeval')
            metric = data.get('metric')
            test_cases_data = data.get('test_cases', [])
            # 按所选指标校验必填字段，缺失则拒绝保存
            validation_errors = TestCaseService.validate_test_cases_for_metric(
                test_cases_data, evaluation_tool, metric
            )
            if validation_errors:
                return {
                    'success': False,
                    'message': '测试项字段校验未通过：\n' + '\n'.join(f'· {e}' for e in validation_errors)
                }

            resolved_agent_id = TestCaseService._resolve_agent_id(user_id, data.get('agent_id'))
            evaluation_set = EvaluationSet(
                user_id=user_id,
                agent_id=resolved_agent_id,
                name=data.get('name'),
                evaluation_tool=evaluation_tool,
                status=data.get('status', 'pending')
            )
            db.session.add(evaluation_set)
            db.session.flush()

            for tc in test_cases_data:
                test_case = TestCase(
                    user_id=user_id,
                    agent_id=evaluation_set.agent_id,
                    set_id=evaluation_set.id,
                    name=tc.get('name', f'测试用例_{len(evaluation_set.test_cases)+1}'),
                    query=tc.get('query'),
                    expected=tc.get('expected'),
                    tags=tc.get('tags', ''),
                    evaluation_tool=evaluation_set.evaluation_tool,
                    metric=tc.get('metric') or metric,
                    input_payload=tc.get('input_payload'),
                    expected_payload=tc.get('expected_payload')
                )
                db.session.add(test_case)

            db.session.commit()
            return {'success': True, 'message': '评测集创建成功', 'data': {'id': evaluation_set.id}}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def materialize_evaluation_set(user_id, virtual_id, data):
        try:
            parts = virtual_id.split('::', 1)
            base_id = parts[0]
            metric_text = parts[1] if len(parts) == 2 else 'none'
            base_parts = base_id.split('-', 2)
            if len(base_parts) != 3 or base_parts[0] != 'orphan':
                return {'success': False, 'message': '评测集不存在'}

            agent_id_text, evaluation_tool = base_parts[1], base_parts[2]
            query = db.session.query(TestCase).filter_by(user_id=user_id, set_id=None, evaluation_tool=evaluation_tool)
            query = query.filter(TestCase.metric.is_(None) if metric_text == 'none' else TestCase.metric == metric_text)
            if agent_id_text == 'none':
                query = query.filter(TestCase.agent_id.is_(None))
                agent_id = data.get('agent_id')
            else:
                agent_id = int(agent_id_text)
                query = query.filter_by(agent_id=agent_id)

            resolved_agent_id = TestCaseService._resolve_agent_id(user_id, data.get('agent_id') or agent_id)
            cases = query.all()
            if not cases:
                return {'success': False, 'message': '默认评测集没有可编辑的测试项'}

            evaluation_set = EvaluationSet(
                user_id=user_id,
                agent_id=resolved_agent_id,
                name=data.get('name'),
                evaluation_tool=data.get('evaluation_tool', evaluation_tool),
                status=data.get('status', 'pending')
            )
            db.session.add(evaluation_set)
            db.session.flush()

            submitted_cases = data.get('test_cases', [])
            for index, case in enumerate(cases):
                submitted = submitted_cases[index] if index < len(submitted_cases) else {}
                case.set_id = evaluation_set.id
                case.agent_id = evaluation_set.agent_id
                case.evaluation_tool = evaluation_set.evaluation_tool
                case.name = submitted.get('name', case.name)
                case.query = submitted.get('query', case.query)
                case.expected = submitted.get('expected', case.expected)
                case.tags = submitted.get('tags', case.tags)
                case.metric = submitted.get('metric') or data.get('metric') or case.metric
                if 'input_payload' in submitted:
                    case.input_payload = submitted.get('input_payload')
                if 'expected_payload' in submitted:
                    case.expected_payload = submitted.get('expected_payload')

            for submitted in submitted_cases[len(cases):]:
                db.session.add(TestCase(
                    user_id=user_id,
                    agent_id=evaluation_set.agent_id,
                    set_id=evaluation_set.id,
                    name=submitted.get('name'),
                    query=submitted.get('query'),
                    expected=submitted.get('expected'),
                    tags=submitted.get('tags', ''),
                    evaluation_tool=evaluation_set.evaluation_tool,
                    metric=submitted.get('metric') or data.get('metric'),
                    input_payload=submitted.get('input_payload'),
                    expected_payload=submitted.get('expected_payload')
                ))

            db.session.commit()
            return {'success': True, 'message': '评测集保存成功', 'data': {'id': evaluation_set.id}}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def update_evaluation_set(user_id, set_id, data):
        try:
            evaluation_set = EvaluationSet.query.filter_by(id=set_id, user_id=user_id).first()
            if not evaluation_set:
                return {'success': False, 'message': '评测集不存在'}

            # 校验：如果提交了 test_cases，按最终（工具，指标）检查必填字段
            submitted_for_validation = data.get('test_cases')
            if submitted_for_validation is not None:
                effective_tool = data.get('evaluation_tool') or evaluation_set.evaluation_tool
                effective_metric = data.get('metric')
                if not effective_metric:
                    # 取任一现有用例的 metric 作参照（同一评测集通常同 metric）
                    for c in evaluation_set.test_cases:
                        if c.metric:
                            effective_metric = c.metric
                            break
                validation_errors = TestCaseService.validate_test_cases_for_metric(
                    submitted_for_validation, effective_tool, effective_metric
                )
                if validation_errors:
                    return {
                        'success': False,
                        'message': '测试项字段校验未通过：\n' + '\n'.join(f'· {e}' for e in validation_errors)
                    }

            if 'name' in data:
                evaluation_set.name = data['name']
            if 'agent_id' in data:
                evaluation_set.agent_id = data['agent_id']
            if 'evaluation_tool' in data:
                evaluation_set.evaluation_tool = data['evaluation_tool']
            if 'status' in data:
                evaluation_set.status = data['status']

            submitted_cases = data.get('test_cases')
            existing_cases = list(evaluation_set.test_cases)
            if submitted_cases is not None:
                for index, submitted in enumerate(submitted_cases):
                    if index < len(existing_cases):
                        case = existing_cases[index]
                        case.agent_id = evaluation_set.agent_id
                        case.evaluation_tool = evaluation_set.evaluation_tool
                        case.name = submitted.get('name', case.name)
                        case.query = submitted.get('query', case.query)
                        case.expected = submitted.get('expected', case.expected)
                        case.tags = submitted.get('tags', case.tags)
                        case.metric = submitted.get('metric') or data.get('metric') or case.metric
                        if 'input_payload' in submitted:
                            case.input_payload = submitted.get('input_payload')
                        if 'expected_payload' in submitted:
                            case.expected_payload = submitted.get('expected_payload')
                    else:
                        db.session.add(TestCase(
                            user_id=user_id,
                            agent_id=evaluation_set.agent_id,
                            set_id=evaluation_set.id,
                            name=submitted.get('name'),
                            query=submitted.get('query'),
                            expected=submitted.get('expected'),
                            tags=submitted.get('tags', ''),
                            evaluation_tool=evaluation_set.evaluation_tool,
                            metric=submitted.get('metric') or data.get('metric'),
                            input_payload=submitted.get('input_payload'),
                            expected_payload=submitted.get('expected_payload')
                        ))

                for case in existing_cases[len(submitted_cases):]:
                    task_cases = TaskTestCase.query.filter_by(test_case_id=case.id).all()
                    for task_case in task_cases:
                        EvaluationResult.query.filter_by(task_case_id=task_case.id).delete()
                        db.session.delete(task_case)
                    db.session.delete(case)
            else:
                for case in existing_cases:
                    case.agent_id = evaluation_set.agent_id
                    case.evaluation_tool = evaluation_set.evaluation_tool

            db.session.commit()
            return {'success': True, 'message': '评测集更新成功'}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def _delete_test_case_with_results(case):
        task_cases = TaskTestCase.query.filter_by(test_case_id=case.id).all()
        for task_case in task_cases:
            EvaluationResult.query.filter_by(task_case_id=task_case.id).delete()
            db.session.delete(task_case)
        db.session.delete(case)

    @staticmethod
    def delete_evaluation_set(user_id, set_id):
        try:
            set_id_text = str(set_id)

            if set_id_text.startswith('orphan-'):
                parts = set_id_text.split('::', 1)
                base_id = parts[0]
                metric_text = parts[1] if len(parts) == 2 else 'none'
                base_parts = base_id.split('-', 2)
                if len(base_parts) != 3:
                    return {'success': False, 'message': '评测集不存在'}

                agent_id_text, evaluation_tool = base_parts[1], base_parts[2]
                query = db.session.query(TestCase).filter_by(user_id=user_id, set_id=None, evaluation_tool=evaluation_tool)
                query = query.filter(TestCase.metric.is_(None) if metric_text == 'none' else TestCase.metric == metric_text)
                if agent_id_text == 'none':
                    query = query.filter(TestCase.agent_id.is_(None))
                else:
                    query = query.filter_by(agent_id=int(agent_id_text))

                cases = query.all()
                if not cases:
                    return {'success': False, 'message': '评测集不存在'}
                for case in cases:
                    TestCaseService._delete_test_case_with_results(case)
                db.session.commit()
                return {'success': True, 'message': '评测集删除成功'}

            evaluation_set = EvaluationSet.query.filter_by(id=int(set_id_text), user_id=user_id).first()
            if not evaluation_set:
                return {'success': False, 'message': '评测集不存在'}

            for case in list(evaluation_set.test_cases):
                TestCaseService._delete_test_case_with_results(case)
            db.session.delete(evaluation_set)
            db.session.commit()
            return {'success': True, 'message': '评测集删除成功'}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def copy_evaluation_set(user_id, set_id):
        try:
            source = EvaluationSet.query.filter_by(id=set_id, user_id=user_id).first()
            if not source:
                return {'success': False, 'message': '评测集不存在'}

            copied_set = EvaluationSet(
                user_id=user_id,
                agent_id=source.agent_id,
                name=f'{source.name} 副本',
                evaluation_tool=source.evaluation_tool,
                status='pending'
            )
            db.session.add(copied_set)
            db.session.flush()

            for source_case in source.test_cases:
                db.session.add(TestCase(
                    user_id=user_id,
                    agent_id=source_case.agent_id,
                    set_id=copied_set.id,
                    name=source_case.name,
                    query=source_case.query,
                    expected=source_case.expected,
                    tags=source_case.tags,
                    evaluation_tool=source_case.evaluation_tool,
                    metric=source_case.metric,
                    input_payload=source_case.input_payload,
                    expected_payload=source_case.expected_payload
                ))

            db.session.commit()
            return {'success': True, 'message': '评测集复制成功', 'data': {'id': copied_set.id}}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}
